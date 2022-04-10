import openpyxl

#엑셀 만들기
wb = openpyxl.Workbook()

#엑셀 워크시트 만들기
ws=wb.create_sheet('june')

#데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'
ws['A2'] = 1
ws['B2'] = '오일남'

#엑셀 저장
wb.save(r'C:\Users\jangd\OneDrive\바탕 화면\Bigdata_process\python\crawling\참가자_data.xlsx')
