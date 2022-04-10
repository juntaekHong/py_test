import openpyxl

fpath = r'C:\Users\jangd\OneDrive\바탕 화면\Bigdata_process\python\crawling\참가자_data.xlsx'
# 엑셀파일 불러오기
wb=openpyxl.load_workbook(fpath)

#엑셀 시트선택
ws=wb['june']

#데이터 수정하기
ws['A3'] = 456
ws['B3']='성기훈'

#엑셀 저장하기

wb.save(fpath)